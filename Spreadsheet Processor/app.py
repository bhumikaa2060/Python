import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_workbook(filename):
    wb = xl.load_workbook(filename)  # loads the workbook
    sheet = wb['Sheet1']  # selects sheet 1

    # to print the third row except for the heading
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        print(cell.value)

    # To increase third value by 0.9 and put it in new column
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    # to make a bar graph:
    values = Reference(sheet,
                       min_row=2,
                       max_row=sheet.max_row,
                       min_col=4,
                       max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, "e2")  # chart appears at e2

    wb.save(filename)  # saves in new file



