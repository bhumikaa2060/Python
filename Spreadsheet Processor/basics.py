import openpyxl as xl

wb = xl.load_workbook('transactions.xlsx') #loads the workbook
sheet = wb['Sheet1'] #selects sheet 1


cell = sheet['a1'] #accesing cell value
sheet.cell(1,1) #accesing cell value
print(cell.value)
print(sheet.max_row) #total number of rows in the sheet

wb.save('transactions2.xlsx')  # saves in new file